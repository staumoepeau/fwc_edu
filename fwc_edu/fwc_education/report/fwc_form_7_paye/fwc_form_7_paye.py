# Copyright (c) 2013, Sione Taumoepeau and contributors
# For license information, please see license.txt
from __future__ import unicode_literals
import json
import time
import math
import ast
import os.path
import sys
import datetime
import pandas as pd
import xlsxwriter
import openpyxl 
import numpy as np

from openpyxl import load_workbook
from frappe import _, msgprint, utils
from datetime import datetime, timedelta
from frappe.utils import flt, getdate, datetime, comma_and
from collections import defaultdict
from werkzeug.wrappers import Response
import frappe, erpnext
from collections import defaultdict


def execute(filters=None):
	data = get_data(filters)
	columns = get_columns(filters) if len(data) else []

	return columns, data

def get_columns(filters):
	columns = [
		{
			"label": _("Employee Name"),
			"options": "Employee",
			"fieldname": "employee_name",
			"fieldtype": "Link",
			"width": 160
		},
		{
			"label": _("TIN"),
			"fieldname": "tin",
			"fieldtype": "Data",
			"width": 140
		},
		{
			"label": _("Income Tax Amount"),
			"fieldname": "it_amount",
			"fieldtype": "Currency",
			"options": "currency",
			"width": 140
		},
		{
			"label": _("Gross Pay"),
			"fieldname": "gross_pay",
			"fieldtype": "Currency",
			"options": "currency",
			"width": 140
		},
	]

	return columns

def get_conditions(filters):
	conditions = [""]

	if filters.get("branch"):
		conditions.append("branch = '%s' " % (filters["branch"]) )

	if filters.get("company"):
		conditions.append("company = '%s' " % (filters["company"]) )

	if filters.get("month"):
		conditions.append("month(posting_date) = '%s' " % (filters["month"]))

	if filters.get("year"):
		conditions.append("year(posting_date) = '%s' " % (filters["year"]))

	return " and ".join(conditions)


def get_data(filters):

	get_tax = {}
	data = []
	taxdata = []

	fields = ["employee", "tin", "company"]

	employee_details = frappe.get_list("Employee", fields = fields)
	employee_data_dict = {}

	for d in employee_details:
		employee_data_dict.setdefault(
			d.employee,{
				"employee" : d.employee,
				"tin": d.tin,
				"company": d.company
			}
		)

	conditions = get_conditions(filters)

	posting_month = filters.get("month")
	posting_year = filters.get("year")
	company = filters.get("company")
	
	get_gross = frappe.db.sql("""SELECT temp.employee, 
			Concat(Ifnull(temp.last_name,' ') ,' ', Ifnull(temp.middle_name,' '),' ', Ifnull(temp.first_name,' ')) as employee_name,
			sal.company, sum(sal.gross_pay) as gross_pay
			FROM `tabEmployee` temp INNER JOIN `tabSalary Slip` sal
			ON temp.employee = sal.employee
			AND month(sal.posting_date) = %s
			AND year(sal.posting_date) = %s
			AND sal.company = %s
			GROUP BY temp.employee
			""", (posting_month, posting_year, company), as_dict=1)
	
	employee = {}

	for e in get_gross:
		employee = {
			"emp_id": e.employee,
			"employee_name" : e.employee_name,
			"tin" : employee_data_dict.get(e.employee).get("tin"),
			"it_amount" : 0,
			"gross_pay": e.gross_pay,
			"company": e.company
		}
		data.append(employee)
	
	get_tax = frappe.db.sql("""SELECT sal.employee, 
			sal.employee_name, 
			sum(ded.amount) as "tax"
			FROM `tabSalary Slip` sal
			INNER JOIN `tabSalary Detail` ded ON
			sal.name = ded.parent
			AND ded.parentfield = 'deductions'
			AND ded.parenttype = 'Salary Slip'
			AND ded.salary_component = "PAYE-TAX"
			AND ded.docstatus = 1
			AND ded.amount > 0
			AND month(sal.posting_date) = %s
			AND year(sal.posting_date) = %s
			AND sal.company = %s
			GROUP BY sal.employee
			""", (posting_month, posting_year, company), as_dict=1)


	for e in get_tax:
		employee = {
			"emp_id": e.employee,
			"employee_name" : e.employee_name,
#			"tin" : 0,
			"it_amount" : e.tax,
#			"gross_pay": 0,
			"company": e.company
		}

		data.append(employee)

	combined = defaultdict(dict)
	
	for item in data:
		combined[item['emp_id']].update(item)
	
	taxdata = (list(combined.values()))

	return taxdata

def get_paye_data(posting_month, posting_year, company):
	data = []
	fields = ["employee", "tin", "company"]
	Month = datetime.date(1900, int(posting_month), 1).strftime('%B')

	employee_details = frappe.get_list("Employee", fields = fields)
	employee_data_dict = {}

	for d in employee_details:
		employee_data_dict.setdefault(
			d.employee,{
				"tin": d.tin,
				"employee" : d.employee,
				"date_of_payment" : '',
				"pay_period": 'Fortnightly',
				"company": d.company
			}
		)

	get_gross = frappe.db.sql("""SELECT temp.employee, 
			Concat(Ifnull(temp.last_name,' ') ,' ', Ifnull(temp.middle_name,' '),' ', Ifnull(temp.first_name,' ')) as employee_name,
			sal.company, sum(sal.gross_pay) as gross_pay
			FROM `tabEmployee` temp INNER JOIN `tabSalary Slip` sal
			ON temp.employee = sal.employee
			AND month(sal.posting_date) = %s
			AND year(sal.posting_date) = %s
			AND sal.company = %s
			GROUP BY temp.employee
			""", (posting_month, posting_year, company), as_dict=1)
	
	employee = {}

	for e in get_gross:
		employee = {
			"emp_id": e.employee,
			"tin" : employee_data_dict.get(e.employee).get("tin"),
			"employee_name" : e.employee_name,
			"date_of_payment" : '',
			"pay_period": 'Fortnightly',
			"gross_pay": e.gross_pay,
			"benefit" : 0,
			"it_amount" : 0,
		}
		data.append(employee)
	
	get_tax = frappe.db.sql("""SELECT sal.employee, 
			sal.employee_name, 
			sum(ded.amount) as "tax"
			FROM `tabSalary Slip` sal
			INNER JOIN `tabSalary Detail` ded ON
			sal.name = ded.parent
			AND ded.parentfield = 'deductions'
			AND ded.parenttype = 'Salary Slip'
			AND ded.salary_component = "PAYE-TAX"
			AND ded.docstatus = 1
			AND ded.amount > 0
			AND month(sal.posting_date) = %s
			AND year(sal.posting_date) = %s
			AND sal.company = %s
			GROUP BY sal.employee
			""", (posting_month, posting_year, company), as_dict=1)


	for e in get_tax:
		employee = {
			"emp_id": e.employee,
#			"employee_name" : e.employee_name,
			"date_of_payment" : '',
			"pay_period": "Fortnightly",
			"it_amount" : e.tax
		}
	
		data.append(employee)
#	frappe.msgprint(_("TAX {0}").format(get_tax))
	combined = defaultdict(dict)
	
	for item in data:
		combined[item['emp_id']].update(item)
	
	taxdata = (list(combined.values()))
#	frappe.msgprint(_("TAX {0}").format(taxdata))
	return taxdata


@frappe.whitelist()
def save_data_to_Excel(month, company, year):

	filename = "PAYE.xlsm"
	Month = datetime.date(1900, int(month), 1).strftime('%B')
	Abbr = frappe.db.get_value("Company", company, "abbr")

	new_filename = Abbr +"-PAYE-" + Month + "-" + year+".xls"

	save_path = 'edu.fwc.to/private/files/'
	file_name = os.path.join(save_path, filename)
	new_file_name = os.path.join(save_path, new_filename)

	ferp = frappe.new_doc("File")
	ferp.file_name = new_filename
	ferp.folder = "Home/PAYE_TAX"
	ferp.is_private = 0
	ferp.file_url = "/private/files/PAYE_TAX/"+new_filename

#	ferp.folder = "Home"
#	ferp.is_private = 1
#	ferp.file_url = "/private/files/Form_7/"+new_filename

	paye_data = []
	paye_data = get_paye_data(month, year, company)

	df = pd.DataFrame(paye_data)

	df = df.drop('emp_id', axis=1)

#	frappe.msgprint(_("Data {0}").format(df))
	workbook1 = openpyxl.load_workbook(file_name, read_only=False, keep_vba= True)
	workbook1.template = True
	sheetname = workbook1.get_sheet_by_name('PAYE')
	sheetname['C5']= str('263317')
	sheetname['C7']= str(Month)
	sheetname['E7']= str(year)
	sheetname['C9']= str('FWC Education')

	writer = pd.ExcelWriter(new_file_name, engine='openpyxl')
	writer.book = workbook1
	writer.sheets = dict((ws.title, ws) for ws in workbook1.worksheets)

	df.to_excel(writer, sheet_name='PAYE', index=False, header=False, startrow=12, startcol=2)

#	with pd.ExcelWriter(file_name) as writer:
#		writer.book = openpyxl.load_workbook(file_name, read_only=False, keep_vba= True)
#		df.to_excel(writer, sheet_name=sheetname, index=False, header=False, startrow=13, startcol=1)

	writer.save()
	ferp.save()
	frappe.db.sql('''UPDATE `tabFile` SET file_url = %s WHERE file_name = %s''',("/files/"+new_filename, new_filename), as_dict=True)
	frappe.msgprint(_("File created - {0}").format(new_filename))
#	frappe.msgprint(_("Form 7 have been created"))
	
#	writer.close()
#	frappe.msgprint(_("Executing the below:"))
#	frappe.local.response.filename = new_file_name
#	with open(new_file_name, "rb") as fileobj:
#		filedata = fileobj.read()
#	frappe.logger().debug("Inside") 
#	frappe.local.response.filecontent = filedata
#	frappe.local.response.type = "download"